from __future__ import annotations

from collections import OrderedDict
from datetime import datetime
from pathlib import Path
import re
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SOURCE_SHEET_NAME = "AFLS Quote"
OUTPUT_SHEET_NAME = "Maersk Rates"
SOURCE_START_COLUMN = 1
SOURCE_END_COLUMN = 16


def sanitize_token(value: object) -> str:
    if value is None:
        return ""

    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", str(value).strip().upper())
    return cleaned.strip("_")


def select_source_sheet(workbook):
    if SOURCE_SHEET_NAME in workbook.sheetnames:
        return workbook[SOURCE_SHEET_NAME]

    for sheet_name in workbook.sheetnames:
        if "QUOTE" in sheet_name.upper():
            return workbook[sheet_name]

    raise ValueError("No quote sheet was found in the uploaded workbook.")


def find_header_row(worksheet) -> int:
    for row_index in range(1, worksheet.max_row + 1):
        charge_value = worksheet.cell(row_index, 12).value
        first_header = worksheet.cell(row_index, 1).value
        if sanitize_token(charge_value) == "CHARGE" and sanitize_token(first_header) == "RECEIPT":
            return row_index

    raise ValueError("Could not locate the AFLS Quote header row.")


def extract_headers(worksheet, header_row: int) -> list[str]:
    headers: list[str] = []
    for column_index in range(SOURCE_START_COLUMN, SOURCE_END_COLUMN + 1):
        raw_header = worksheet.cell(header_row, column_index).value
        headers.append(str(raw_header).strip() if raw_header is not None else f"Column {column_index}")
    return headers


def get_column_map(headers: Iterable[str]) -> dict[str, int]:
    header_list = list(headers)
    normalized_to_index: dict[str, int] = {}
    for index, header in enumerate(header_list, start=1):
        token = sanitize_token(header)
        if token:
            normalized_to_index[token] = index
    return normalized_to_index


def extract_container_headers(headers: Iterable[str], column_map: dict[str, int]) -> dict[int, str]:
    header_list = list(headers)
    container_indexes = [
        column_map["20DRY"],
        column_map["40DRY"],
        column_map["40HDRY"],
    ]
    return {column_index: sanitize_token(header_list[column_index - 1]) for column_index in container_indexes}


def parse_amount_and_currency(value: object) -> tuple[object, str]:
    if value is None:
        return "", "UNKNOWN"

    text = str(value).strip()
    match = re.match(r"^\s*([-+]?\d+(?:\.\d+)?)\s*([A-Za-z]{3})\s*$", text)
    if not match:
        return text, "UNKNOWN"

    amount_text, currency = match.groups()
    amount = float(amount_text) if "." in amount_text else int(amount_text)
    return amount, currency.upper()


def collect_groups(worksheet, header_row: int):
    source_headers = extract_headers(worksheet, header_row)
    column_map = get_column_map(source_headers)
    charge_column_index = column_map["CHARGE"]
    rate_basis_column_index = column_map["RATE_BASIS"]
    container_headers = extract_container_headers(source_headers, column_map)
    dynamic_headers: "OrderedDict[str, None]" = OrderedDict()
    charge_currency_map: "OrderedDict[tuple[str, str], OrderedDict[str, None]]" = OrderedDict()
    output_rows: list[dict[str, object]] = []
    current_row: dict[str, object] | None = None

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        row_values = [
            worksheet.cell(row_index, column_index).value
            for column_index in range(SOURCE_START_COLUMN, SOURCE_END_COLUMN + 1)
        ]

        if not any(value is not None and str(value).strip() != "" for value in row_values):
            continue

        charge_code = sanitize_token(row_values[charge_column_index - 1])
        if not charge_code:
            continue

        if charge_code == "BAS":
            if current_row is not None:
                output_rows.append(current_row)

            current_row = {
                "base_values": row_values,
                "charges": {},
            }
            continue

        if current_row is None:
            continue

        rate_basis = sanitize_token(row_values[rate_basis_column_index - 1]) or "RATE"
        charges: dict[str, object] = current_row["charges"]  # type: ignore[assignment]
        charge_key = (charge_code, rate_basis)
        charge_currency_map.setdefault(charge_key, OrderedDict())

        for column_index in container_headers:
            value = row_values[column_index - 1]
            if value in (None, ""):
                continue

            amount, currency = parse_amount_and_currency(value)
            charge_currency_map[charge_key].setdefault(currency, None)
            header_name = f"{charge_code}_{container_headers[column_index]}_{rate_basis}_{currency}"

            existing_value = charges.get(header_name)
            if existing_value not in (None, "") and existing_value != amount:
                charges[header_name] = f"{existing_value} | {amount}"
            else:
                charges[header_name] = amount

    for (charge_code, rate_basis), currencies in charge_currency_map.items():
        for currency in currencies:
            for column_index in container_headers:
                header_name = f"{charge_code}_{container_headers[column_index]}_{rate_basis}_{currency}"
                dynamic_headers.setdefault(header_name, None)

    if current_row is not None:
        output_rows.append(current_row)

    return source_headers, list(dynamic_headers.keys()), output_rows, column_map


def style_output_sheet(worksheet, total_columns: int, total_rows: int) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="24305E")
    accent_fill = PatternFill(fill_type="solid", fgColor="141B34")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="30416F"),
        right=Side(style="thin", color="30416F"),
        top=Side(style="thin", color="30416F"),
        bottom=Side(style="thin", color="30416F"),
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in worksheet.iter_rows(min_row=2, max_row=total_rows, min_col=1, max_col=total_columns):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for column_index in range(17, total_columns + 1):
        worksheet.cell(1, column_index).fill = accent_fill

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 30)

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.auto_filter.ref = f"A1:{get_column_letter(total_columns)}{total_rows}"


def build_processed_workbook(source_workbook):
    source_sheet = select_source_sheet(source_workbook)
    header_row = find_header_row(source_sheet)
    source_headers, dynamic_headers, grouped_rows, column_map = collect_groups(source_sheet, header_row)
    workbook = Workbook()
    output_sheet = workbook.active
    output_sheet.title = OUTPUT_SHEET_NAME
    output_sheet.append(source_headers + dynamic_headers)

    for grouped_row in grouped_rows:
        base_values = grouped_row["base_values"]
        charge_values = grouped_row["charges"]
        row = list(base_values) + [charge_values.get(header_name, "") for header_name in dynamic_headers]
        output_sheet.append(row)

    style_output_sheet(
        output_sheet,
        total_columns=len(source_headers) + len(dynamic_headers),
        total_rows=max(len(grouped_rows) + 1, 1),
    )

    effective_date_column_index = column_map.get("EFFECTIVE_DATE")
    expiry_date_column_index = column_map.get("EXPIRY_DATE")
    if effective_date_column_index and expiry_date_column_index:
        for row_index in range(2, output_sheet.max_row + 1):
            output_sheet.cell(row_index, effective_date_column_index).number_format = "yyyy-mm-dd"
            output_sheet.cell(row_index, expiry_date_column_index).number_format = "yyyy-mm-dd"

    return workbook


def process_workbook(input_path: str | Path, output_path: str | Path) -> Path:
    input_file = Path(input_path)
    output_file = Path(output_path)
    source_workbook = load_workbook(input_file)
    workbook = build_processed_workbook(source_workbook)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)
    return output_file
